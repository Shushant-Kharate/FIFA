from pathlib import Path
from io import BytesIO

import pytest
from openpyxl import load_workbook
from fastapi.testclient import TestClient
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker
from sqlalchemy.pool import StaticPool

import models
from database import Base, get_db
from main import app
from services.excel_import import process_excel_import

TEST_ROOM1_PASSWORD = "test-room1-password"
TEST_ROOM2_PASSWORD = "test-room2-password"
TEST_SUPER_ADMIN_PASSWORD = "test-super-admin-password"


@pytest.fixture()
def api_context(monkeypatch):
    monkeypatch.setenv("AUTH_SECRET", "test-auth-secret-with-more-than-32-characters")
    monkeypatch.setenv("ROOM1_ADMIN_PASSWORD", TEST_ROOM1_PASSWORD)
    monkeypatch.setenv("ROOM2_ADMIN_PASSWORD", TEST_ROOM2_PASSWORD)
    monkeypatch.setenv("SUPER_ADMIN_PASSWORD", TEST_SUPER_ADMIN_PASSWORD)

    engine = create_engine(
        "sqlite://",
        connect_args={"check_same_thread": False},
        poolclass=StaticPool,
    )
    testing_session = sessionmaker(autocommit=False, autoflush=False, bind=engine)
    Base.metadata.create_all(bind=engine)
    seed_db = testing_session()
    seed_db.add_all([models.Room(id=1, name="Room 1"), models.Room(id=2, name="Room 2")])
    seed_db.commit()
    workbook = Path(__file__).with_name("FIFA AUCTION 2026.xlsx").read_bytes()
    assert process_excel_import(workbook, "FIFA AUCTION 2026.xlsx", seed_db, 1).success
    assert process_excel_import(workbook, "FIFA AUCTION 2026.xlsx", seed_db, 2).success
    seed_db.close()

    def override_get_db():
        db = testing_session()
        try:
            yield db
        finally:
            db.close()

    app.dependency_overrides[get_db] = override_get_db
    client = TestClient(app)
    yield client, testing_session, workbook
    app.dependency_overrides.clear()
    engine.dispose()


def _login(client, username, password):
    response = client.post("/api/auth/login", json={"username": username, "password": password})
    assert response.status_code == 200, response.text
    return response.json()["access_token"]


def _headers(token, room_id=None):
    result = {"Authorization": f"Bearer {token}"}
    if room_id is not None:
        result["X-Room-ID"] = str(room_id)
    return result


def test_all_business_routes_and_room_isolation(api_context):
    client, testing_session, workbook = api_context

    assert client.get("/api/health").json()["database"] == "online"
    assert client.get("/api/players").status_code == 401
    assert client.post(
        "/api/auth/login", json={"username": "room1_admin", "password": "wrong"}
    ).status_code == 401

    room1_token = _login(client, "room1_admin", TEST_ROOM1_PASSWORD)
    room2_token = _login(client, "room2_admin", TEST_ROOM2_PASSWORD)
    super_token = _login(client, "super_admin", TEST_SUPER_ADMIN_PASSWORD)
    room1_headers = _headers(room1_token, 2)  # Room admins cannot override their room.
    room2_headers = _headers(room2_token, 1)
    super_room2 = _headers(super_token, 2)

    assert client.get("/api/auth/me", headers=room1_headers).json()["room_id"] == 1
    room1_players = client.get(
        "/api/players?position=GK&status=AVAILABLE&sort=score_desc", headers=room1_headers
    ).json()
    room2_players = client.get("/api/players", headers=room2_headers).json()
    assert room1_players and len(room2_players) == 152
    first_player_id = room1_players[0]["id"]
    assert client.get(f"/api/players/{first_player_id}", headers=room1_headers).status_code == 200
    assert client.get(f"/api/players/{first_player_id}", headers=room2_headers).status_code == 404

    room1_teams = client.get("/api/teams", headers=room1_headers).json()
    room2_teams = client.get("/api/teams", headers=room2_headers).json()
    assert len(room1_teams) == len(room2_teams) == 20
    room1_team_id = room1_teams[0]["team_id"]
    assert client.get(f"/api/teams/{room1_team_id}", headers=room1_headers).status_code == 200
    assert client.get("/api/results", headers=room1_headers).status_code == 200

    assert client.put(
        "/api/settings", headers=room1_headers, json={"starting_budget": "800"}
    ).json()["starting_budget"] == "800"
    assert client.get("/api/settings", headers=room2_headers).json()["starting_budget"] == "700"
    assert client.put(
        "/api/settings", headers=room1_headers, json={"starting_budget": "0"}
    ).status_code == 422
    assert client.put(
        "/api/settings", headers=room1_headers, json={"starting_budget": "NaN"}
    ).status_code == 422
    assert client.put(
        "/api/settings", headers=room1_headers, json={"required_gk": "abc"}
    ).status_code == 422
    assert client.put(
        "/api/settings", headers=room1_headers, json={"unexpected_setting": "1"}
    ).status_code == 422
    assert client.get("/api/teams", headers=room1_headers).status_code == 200

    sell = client.post(
        "/api/auction/sell",
        headers=room1_headers,
        json={"player_id": first_player_id, "team_id": room1_team_id, "price": 5},
    )
    assert sell.status_code == 200 and sell.json()["spent"] == 5
    budget_cut = client.put(
        "/api/settings", headers=room1_headers, json={"starting_budget": "4"}
    )
    assert budget_cut.status_code == 422
    state_after_rejected_cut = client.get(
        f"/api/teams/{room1_team_id}", headers=room1_headers
    ).json()
    assert state_after_rejected_cut["starting_budget"] == 800
    assert state_after_rejected_cut["remaining_budget"] == 795
    assert client.get("/api/auction/history", headers=room1_headers).json()[0]["event_type"] == "SOLD"
    assert client.post(
        "/api/auction/undo", headers=room1_headers, json={"player_id": first_player_id}
    ).json()["status"] == "AVAILABLE"
    assert client.post(
        "/api/auction/undo", headers=room1_headers, json={"player_id": first_player_id}
    ).status_code == 400
    assert client.post(
        "/api/auction/unsold", headers=room1_headers, json={"player_id": first_player_id}
    ).json()["status"] == "UNSOLD"
    assert client.post(
        "/api/auction/return-to-pool", headers=room1_headers, json={"player_id": first_player_id}
    ).json()["status"] == "AVAILABLE"

    # A sold player cannot bypass the guarded sale-undo workflow via UNSOLD.
    assert client.post(
        "/api/auction/sell",
        headers=room1_headers,
        json={"player_id": first_player_id, "team_id": room1_team_id, "price": 5},
    ).status_code == 200
    assert client.post(
        "/api/auction/unsold", headers=room1_headers, json={"player_id": first_player_id}
    ).status_code == 400
    assert client.post(
        "/api/auction/undo", headers=room1_headers, json={"player_id": first_player_id}
    ).status_code == 200

    db = testing_session()
    squad = []
    for position, count in (("GK", 1), ("DEF", 3), ("MID", 2), ("ATT", 2)):
        squad.extend(
            db.query(models.Player).filter(
                models.Player.room_id == 1,
                models.Player.position == position,
                models.Player.status == "AVAILABLE",
            ).limit(count).all()
        )
    squad_ids = [player.id for player in squad]
    db.close()
    assert len(squad_ids) == 8
    for player_id in squad_ids:
        response = client.post(
            "/api/auction/sell",
            headers=room1_headers,
            json={"player_id": player_id, "team_id": room1_team_id, "price": 1},
        )
        assert response.status_code == 200, response.text
    qualified = client.get(f"/api/teams/{room1_team_id}", headers=room1_headers).json()
    assert qualified["qualified"] and len(qualified["best_8_ids"]) == 8
    captain = client.post(
        f"/api/teams/{room1_team_id}/captain",
        headers=room1_headers,
        json={"player_id": qualified["best_8_ids"][0]},
    )
    assert captain.status_code == 200 and captain.json()["captain_id"] is not None

    backup = client.get("/api/admin/backup", headers=room1_headers)
    assert backup.status_code == 200 and backup.json()["room_id"] == 1
    audit_log = client.get("/api/admin/audit-log", headers=room1_headers)
    assert audit_log.status_code == 200
    assert any(entry["action"] == "CAPTAIN_SET" for entry in audit_log.json())
    results_excel = client.get("/api/admin/export-results", headers=room1_headers)
    assert results_excel.status_code == 200 and results_excel.content[:2] == b"PK"
    results_workbook = load_workbook(BytesIO(results_excel.content), read_only=True)
    assert results_workbook.sheetnames == ["Leaderboard", "Team Squads"]
    assert results_workbook["Leaderboard"]["A2"].value == 1
    assert results_workbook["Leaderboard"]["C2"].value == "Yes"
    assert results_workbook["Team Squads"]["E2"].value
    audit_excel = client.get("/api/admin/export-audit-log", headers=room1_headers)
    assert audit_excel.status_code == 200 and audit_excel.content[:2] == b"PK"
    audit_workbook = load_workbook(BytesIO(audit_excel.content), read_only=True)
    assert audit_workbook.sheetnames == ["Audit Log", "Transaction Ledger"]
    assert audit_workbook["Audit Log"].max_row > 1
    template = client.get("/api/admin/sample-template", headers=room1_headers)
    assert template.status_code == 200 and template.content[:2] == b"PK"
    assert client.post(
        "/api/admin/import",
        headers=room1_headers,
        files={"file": ("players.xlsx", workbook)},
    ).status_code == 403
    assert client.post(
        "/api/admin/import",
        headers=super_room2,
        files={"file": ("PLAYERS.XLSX", workbook)},
    ).json()["player_count"] == 152
    # Replacing Room 2's dataset did not change Room 1's qualified squad.
    assert client.get(f"/api/teams/{room1_team_id}", headers=room1_headers).json()["qualified"]
    assert client.post(
        "/api/admin/import",
        headers=super_room2,
        files={"file": ("players.txt", b"not a dataset")},
    ).status_code == 400
    assert client.post(
        "/api/admin/import",
        headers=super_room2,
        files={"file": ("too-large.csv", b"x" * (4 * 1024 * 1024 + 1))},
    ).status_code == 413
    reset = client.post("/api/admin/reset", headers=room2_headers)
    assert reset.status_code == 200 and reset.json()["room_id"] == 2

    expected_business_routes = {
        ("GET", "/api/health"),
        ("POST", "/api/auth/login"),
        ("GET", "/api/auth/me"),
        ("GET", "/api/players"),
        ("GET", "/api/players/{player_id}"),
        ("GET", "/api/teams"),
        ("GET", "/api/teams/{team_id}"),
        ("POST", "/api/teams/{team_id}/captain"),
        ("POST", "/api/auction/sell"),
        ("POST", "/api/auction/unsold"),
        ("POST", "/api/auction/undo"),
        ("POST", "/api/auction/return-to-pool"),
        ("GET", "/api/auction/history"),
        ("GET", "/api/results"),
        ("POST", "/api/admin/import"),
        ("GET", "/api/admin/sample-template"),
            ("GET", "/api/admin/backup"),
            ("GET", "/api/admin/audit-log"),
            ("GET", "/api/admin/export-results"),
            ("GET", "/api/admin/export-audit-log"),
        ("POST", "/api/admin/reset"),
        ("POST", "/api/admin/scale-dataset"),
        ("GET", "/api/admin/removed-players"),
        ("GET", "/api/admin/export-removed-players"),
        ("GET", "/api/settings"),
        ("PUT", "/api/settings"),
    }


    actual_business_routes = {
        (method.upper(), path)
        for path, operations in app.openapi()["paths"].items()
        for method in operations
        if path.startswith("/api/")
    }
    assert actual_business_routes == expected_business_routes
