import io
import json

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session

from models import AuditLog, Player, Team, Transaction
from services.scoring import get_all_teams_leaderboard


HEADER_FILL = PatternFill("solid", fgColor="17365D")
QUALIFIED_FILL = PatternFill("solid", fgColor="E2F0D9")


def _style_sheet(ws, freeze: str = "A2") -> None:
    ws.freeze_panes = freeze
    ws.auto_filter.ref = ws.dimensions
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")
    for column in ws.columns:
        values = [str(cell.value or "") for cell in column]
        width = min(max(max(map(len, values)) + 2, 10), 42)
        ws.column_dimensions[get_column_letter(column[0].column)].width = width


def _save(workbook: Workbook) -> bytes:
    output = io.BytesIO()
    workbook.save(output)
    return output.getvalue()


def generate_results_excel(db: Session, room_id: int) -> bytes:
    leaderboard = get_all_teams_leaderboard(db, room_id)
    workbook = Workbook()
    summary = workbook.active
    summary.title = "Leaderboard"
    summary.append([
        "Rank", "Team", "Qualified", "Final Score", "Base Score", "Captain Bonus",
        "Nationality Bonus", "Club Bonus", "Players", "Spent (€M)", "Remaining (€M)",
    ])
    for rank, team in enumerate(leaderboard, 1):
        summary.append([
            rank, f"Team {team.team_number:02d}", "Yes" if team.qualified else "No",
            team.final_score, team.base_score, team.captain_score,
            team.nationality_bonus, team.club_bonus, len(team.players),
            team.spent, team.remaining_budget,
        ])
        if team.qualified:
            for cell in summary[summary.max_row]:
                cell.fill = QUALIFIED_FILL
    _style_sheet(summary)

    squads = workbook.create_sheet("Team Squads")
    squads.append([
        "Rank", "Team", "Qualified", "Player Code", "Player", "Position",
        "Nationality", "Club", "Player Score", "Bought Price (€M)",
        "Best 8", "Captain", "Team Final Score",
    ])
    for rank, team in enumerate(leaderboard, 1):
        players = sorted(
            team.players,
            key=lambda player: (player.position, -player.score, player.name.lower()),
        )
        if not players:
            squads.append([
                rank, f"Team {team.team_number:02d}", "Yes" if team.qualified else "No",
                "", "No players purchased", "", "", "", "", "", "", "", team.final_score,
            ])
        for player in players:
            squads.append([
                rank, f"Team {team.team_number:02d}", "Yes" if team.qualified else "No",
                player.player_code, player.name, player.position, player.nationality or "",
                player.club or "", player.score, player.sold_price or 0,
                "Yes" if player.id in team.best_8_ids else "No",
                "Yes" if player.is_captain else "No", team.final_score,
            ])
    _style_sheet(squads)
    return _save(workbook)


def generate_audit_excel(db: Session, room_id: int) -> bytes:
    workbook = Workbook()
    audit_sheet = workbook.active
    audit_sheet.title = "Audit Log"
    audit_sheet.append([
        "Log ID", "Timestamp (UTC)", "Action", "Admin", "Description",
        "Player ID", "Team", "Amount (€M)", "Details",
    ])
    team_numbers = {
        team.id: team.team_number
        for team in db.query(Team).filter(Team.room_id == room_id).all()
    }
    entries = db.query(AuditLog).filter(
        AuditLog.room_id == room_id
    ).order_by(AuditLog.timestamp.asc(), AuditLog.id.asc()).all()
    for entry in entries:
        audit_sheet.append([
            entry.id,
            entry.timestamp.isoformat(sep=" ", timespec="seconds") if entry.timestamp else "",
            entry.action,
            entry.actor_username,
            entry.description,
            entry.player_id or "",
            team_numbers.get(entry.team_id, "") if entry.team_id else "",
            entry.amount if entry.amount is not None else "",
            json.dumps(json.loads(entry.details_json), ensure_ascii=False) if entry.details_json else "",
        ])
    _style_sheet(audit_sheet)

    legacy = workbook.create_sheet("Transaction Ledger")
    legacy.append([
        "Transaction ID", "Timestamp (UTC)", "Event", "Player Code", "Player",
        "Team", "Amount (€M)",
    ])
    players = {
        player.id: player
        for player in db.query(Player).filter(Player.room_id == room_id).all()
    }
    transactions = db.query(Transaction).filter(
        Transaction.room_id == room_id
    ).order_by(Transaction.timestamp.asc(), Transaction.id.asc()).all()
    for transaction in transactions:
        player = players.get(transaction.player_id)
        legacy.append([
            transaction.id,
            transaction.timestamp.isoformat(sep=" ", timespec="seconds") if transaction.timestamp else "",
            transaction.event_type,
            player.player_code if player else "",
            player.name if player else "",
            team_numbers.get(transaction.team_id, "") if transaction.team_id else "",
            transaction.amount if transaction.amount is not None else "",
        ])
    _style_sheet(legacy)
    return _save(workbook)


def generate_removed_players_excel(removed_players: list) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Removed Players"
    sheet.append([
        "Player Code", "Name", "Position", "Score", "P1", "P2", "P3",
        "Base Price (€M)", "Nationality", "Club",
    ])
    for p in removed_players:
        sheet.append([
            p.get("player_code", ""),
            p.get("name", ""),
            p.get("position", ""),
            p.get("score", 0),
            p.get("p1", 0),
            p.get("p2", 0),
            p.get("p3", 0),
            p.get("base_price", 1.0),
            p.get("nationality", "") or "",
            p.get("club", "") or "",
        ])
    _style_sheet(sheet)
    return _save(workbook)

